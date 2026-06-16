from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook

from xlsx_agent.core.errors import XlsxAgentError
from xlsx_agent.core.store import WorkbookStore
from xlsx_agent.core.workbook import workbook_close, workbook_open, workbook_save


def _create_workbook(path: Path, *, cell_value: str = "hello") -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet["A1"] = cell_value
    workbook.save(path)
    workbook.close()


def test_workbook_open_returns_handle(tmp_path: Path) -> None:
    workbook_path = tmp_path / "sample.xlsx"
    _create_workbook(workbook_path)

    handle = workbook_open(WorkbookStore(), path=str(workbook_path))

    assert handle.workbook_id.startswith("wb_")
    assert handle.path == str(workbook_path.resolve())
    assert handle.read_only is False
    assert handle.dirty is False


def test_workbook_open_fails_for_missing_file(tmp_path: Path) -> None:
    missing_path = tmp_path / "missing.xlsx"

    try:
        workbook_open(WorkbookStore(), path=str(missing_path))
    except XlsxAgentError as exc:
        assert exc.code == "file_not_found"
    else:
        raise AssertionError("Expected workbook_open to fail for missing file.")


def test_workbook_save_resets_dirty_and_persists_changes(tmp_path: Path) -> None:
    workbook_path = tmp_path / "sample.xlsx"
    _create_workbook(workbook_path, cell_value="before")
    store = WorkbookStore()

    handle = workbook_open(store, path=str(workbook_path))
    session = store.get(handle.workbook_id)
    session.workbook.active["A1"] = "after"
    store.mark_dirty(handle.workbook_id)

    saved = workbook_save(store, workbook_id=handle.workbook_id)
    persisted = load_workbook(workbook_path)

    assert saved.dirty is False
    assert persisted.active["A1"].value == "after"
    persisted.close()


def test_workbook_save_requires_overwrite_for_save_as(tmp_path: Path) -> None:
    source_path = tmp_path / "source.xlsx"
    target_path = tmp_path / "target.xlsx"
    _create_workbook(source_path)
    _create_workbook(target_path)
    store = WorkbookStore()

    handle = workbook_open(store, path=str(source_path))

    try:
        workbook_save(store, workbook_id=handle.workbook_id, path=str(target_path))
    except XlsxAgentError as exc:
        assert exc.code == "overwrite_required"
    else:
        raise AssertionError("Expected save-as without overwrite to fail.")


def test_workbook_save_detects_conflict_on_original_path(tmp_path: Path) -> None:
    workbook_path = tmp_path / "sample.xlsx"
    _create_workbook(workbook_path, cell_value="before")
    store = WorkbookStore()

    handle = workbook_open(store, path=str(workbook_path))
    session = store.get(handle.workbook_id)
    session.workbook.active["A1"] = "after"
    store.mark_dirty(handle.workbook_id)

    replacement = Workbook()
    replacement.active["A1"] = "external"
    replacement.save(workbook_path)
    replacement.close()

    try:
        workbook_save(store, workbook_id=handle.workbook_id)
    except XlsxAgentError as exc:
        assert exc.code == "save_conflict"
    else:
        raise AssertionError("Expected conflicting save to fail.")


def test_workbook_close_requires_force_when_dirty(tmp_path: Path) -> None:
    workbook_path = tmp_path / "sample.xlsx"
    _create_workbook(workbook_path)
    store = WorkbookStore()

    handle = workbook_open(store, path=str(workbook_path))
    store.mark_dirty(handle.workbook_id)

    try:
        workbook_close(store, workbook_id=handle.workbook_id)
    except XlsxAgentError as exc:
        assert exc.code == "invalid_input"
    else:
        raise AssertionError("Expected close without force to fail for dirty workbook.")


def test_workbook_close_removes_session(tmp_path: Path) -> None:
    workbook_path = tmp_path / "sample.xlsx"
    _create_workbook(workbook_path)
    store = WorkbookStore()

    handle = workbook_open(store, path=str(workbook_path))
    closed = workbook_close(store, workbook_id=handle.workbook_id)

    assert closed.workbook_id == handle.workbook_id
    assert store.count() == 0


def test_store_enforces_max_open_workbooks(tmp_path: Path) -> None:
    workbook_path = tmp_path / "sample.xlsx"
    _create_workbook(workbook_path)
    store = WorkbookStore(max_open_workbooks=1)

    first = workbook_open(store, path=str(workbook_path))
    assert first.workbook_id.startswith("wb_")

    try:
        workbook_open(store, path=str(workbook_path))
    except XlsxAgentError as exc:
        assert exc.code == "unsupported_operation"
    else:
        raise AssertionError("Expected store limit to be enforced.")
