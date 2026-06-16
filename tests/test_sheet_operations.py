from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook

from xlsx_agent.core.errors import XlsxAgentError
from xlsx_agent.core.sheets import sheet_create, sheet_list
from xlsx_agent.core.store import WorkbookStore
from xlsx_agent.core.workbook import workbook_open


def _create_workbook(path: Path) -> None:
    workbook = Workbook()
    first = workbook.active
    first.title = "Summary"
    first["A1"] = "name"
    first["B2"] = 10

    hidden = workbook.create_sheet(title="Hidden")
    hidden.sheet_state = "hidden"

    very_hidden = workbook.create_sheet(title="VeryHidden")
    very_hidden.sheet_state = "veryHidden"

    workbook.save(path)
    workbook.close()


def test_sheet_list_returns_basic_metadata(tmp_path: Path) -> None:
    workbook_path = tmp_path / "sheets.xlsx"
    _create_workbook(workbook_path)
    store = WorkbookStore()

    handle = workbook_open(store, path=str(workbook_path))
    sheets = sheet_list(store, workbook_id=handle.workbook_id)

    assert [sheet.name for sheet in sheets] == ["Summary", "Hidden", "VeryHidden"]
    assert sheets[0].position == 1
    assert sheets[0].is_active is True
    assert sheets[0].visibility == "visible"
    assert sheets[0].used_range == "A1:B2"
    assert sheets[1].visibility == "hidden"
    assert sheets[2].visibility == "very_hidden"


def test_sheet_create_adds_sheet_and_marks_workbook_dirty(tmp_path: Path) -> None:
    workbook_path = tmp_path / "create.xlsx"
    _create_workbook(workbook_path)
    store = WorkbookStore()

    handle = workbook_open(store, path=str(workbook_path))
    created = sheet_create(store, workbook_id=handle.workbook_id, name="Report")
    sheets = sheet_list(store, workbook_id=handle.workbook_id)

    assert created.name == "Report"
    assert created.position == 4
    assert store.get(handle.workbook_id).dirty is True
    assert sheets[-1].name == "Report"
    assert sheets[-1].used_range is None


def test_sheet_create_respects_position(tmp_path: Path) -> None:
    workbook_path = tmp_path / "position.xlsx"
    _create_workbook(workbook_path)
    store = WorkbookStore()

    handle = workbook_open(store, path=str(workbook_path))
    created = sheet_create(store, workbook_id=handle.workbook_id, name="Inserted", position=2)
    sheets = sheet_list(store, workbook_id=handle.workbook_id)

    assert created.position == 2
    assert [sheet.name for sheet in sheets] == ["Summary", "Inserted", "Hidden", "VeryHidden"]


def test_sheet_create_rejects_duplicate_names(tmp_path: Path) -> None:
    workbook_path = tmp_path / "duplicate.xlsx"
    _create_workbook(workbook_path)
    store = WorkbookStore()

    handle = workbook_open(store, path=str(workbook_path))

    try:
        sheet_create(store, workbook_id=handle.workbook_id, name="Summary")
    except XlsxAgentError as exc:
        assert exc.code == "invalid_input"
    else:
        raise AssertionError("Expected duplicate sheet name to fail.")


def test_sheet_create_rejects_read_only_workbooks(tmp_path: Path) -> None:
    workbook_path = tmp_path / "readonly.xlsx"
    _create_workbook(workbook_path)
    store = WorkbookStore()

    handle = workbook_open(store, path=str(workbook_path), read_only=True)

    try:
        sheet_create(store, workbook_id=handle.workbook_id, name="Report")
    except XlsxAgentError as exc:
        assert exc.code == "unsupported_operation"
    else:
        raise AssertionError("Expected sheet_create to fail for read-only workbooks.")
