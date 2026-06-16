from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook

from xlsx_agent.core.errors import XlsxAgentError
from xlsx_agent.core.formula import formula_set
from xlsx_agent.core.read import range_read
from xlsx_agent.core.sheets import sheet_create, sheet_delete, sheet_list, sheet_rename
from xlsx_agent.core.store import WorkbookStore
from xlsx_agent.core.workbook import workbook_open
from xlsx_agent.core.write import columns_modify, rows_modify, table_write


def _create_workbook(path: Path) -> None:
    workbook = Workbook()
    first = workbook.active
    first.title = "Sheet1"
    first["A1"] = "Name"
    first["B1"] = "Amount"
    first["A2"] = "Alice"
    first["B2"] = 10

    second = workbook.create_sheet(title="Data")
    second["A1"] = "X"
    second["B1"] = "Y"
    second["A2"] = 1
    second["B2"] = 2

    workbook.save(path)
    workbook.close()


def _create_single_sheet_workbook(path: Path) -> None:
    workbook = Workbook()
    workbook.active.title = "OnlySheet"
    workbook.save(path)
    workbook.close()


# --- sheet_delete ---


def test_sheet_delete_removes_sheet(tmp_path: Path) -> None:
    workbook_path = tmp_path / "delete.xlsx"
    _create_workbook(workbook_path)
    store = WorkbookStore()

    handle = workbook_open(store, path=str(workbook_path))
    result = sheet_delete(store, workbook_id=handle.workbook_id, sheet="Data")
    sheets = sheet_list(store, workbook_id=handle.workbook_id)

    assert result["deleted_sheet"] == "Data"
    assert result["dirty"] is True
    assert len(sheets) == 1
    assert sheets[0].name == "Sheet1"


def test_sheet_delete_rejects_missing_sheet(tmp_path: Path) -> None:
    workbook_path = tmp_path / "delete-missing.xlsx"
    _create_workbook(workbook_path)
    store = WorkbookStore()

    handle = workbook_open(store, path=str(workbook_path))

    try:
        sheet_delete(store, workbook_id=handle.workbook_id, sheet="Nonexistent")
    except XlsxAgentError as exc:
        assert exc.code == "sheet_not_found"
    else:
        raise AssertionError("Expected delete of missing sheet to fail.")


def test_sheet_delete_rejects_last_sheet(tmp_path: Path) -> None:
    workbook_path = tmp_path / "delete-last.xlsx"
    _create_single_sheet_workbook(workbook_path)
    store = WorkbookStore()

    handle = workbook_open(store, path=str(workbook_path))

    try:
        sheet_delete(store, workbook_id=handle.workbook_id, sheet="OnlySheet")
    except XlsxAgentError as exc:
        assert exc.code == "invalid_input"
    else:
        raise AssertionError("Expected delete of last sheet to fail.")


def test_sheet_delete_rejects_read_only(tmp_path: Path) -> None:
    workbook_path = tmp_path / "delete-ro.xlsx"
    _create_workbook(workbook_path)
    store = WorkbookStore()

    handle = workbook_open(store, path=str(workbook_path), read_only=True)

    try:
        sheet_delete(store, workbook_id=handle.workbook_id, sheet="Data")
    except XlsxAgentError as exc:
        assert exc.code == "unsupported_operation"
    else:
        raise AssertionError("Expected delete on read-only workbook to fail.")


# --- sheet_rename ---


def test_sheet_rename_changes_name(tmp_path: Path) -> None:
    workbook_path = tmp_path / "rename.xlsx"
    _create_workbook(workbook_path)
    store = WorkbookStore()

    handle = workbook_open(store, path=str(workbook_path))
    result = sheet_rename(store, workbook_id=handle.workbook_id, sheet="Sheet1", new_name="Summary")
    sheets = sheet_list(store, workbook_id=handle.workbook_id)

    assert result["old_name"] == "Sheet1"
    assert result["new_name"] == "Summary"
    assert result["dirty"] is True
    assert sheets[0].name == "Summary"


def test_sheet_rename_rejects_missing_sheet(tmp_path: Path) -> None:
    workbook_path = tmp_path / "rename-missing.xlsx"
    _create_workbook(workbook_path)
    store = WorkbookStore()

    handle = workbook_open(store, path=str(workbook_path))

    try:
        sheet_rename(store, workbook_id=handle.workbook_id, sheet="Nonexistent", new_name="New")
    except XlsxAgentError as exc:
        assert exc.code == "sheet_not_found"
    else:
        raise AssertionError("Expected rename of missing sheet to fail.")


def test_sheet_rename_rejects_duplicate_name(tmp_path: Path) -> None:
    workbook_path = tmp_path / "rename-dup.xlsx"
    _create_workbook(workbook_path)
    store = WorkbookStore()

    handle = workbook_open(store, path=str(workbook_path))

    try:
        sheet_rename(store, workbook_id=handle.workbook_id, sheet="Sheet1", new_name="Data")
    except XlsxAgentError as exc:
        assert exc.code == "invalid_input"
    else:
        raise AssertionError("Expected rename to duplicate name to fail.")


def test_sheet_rename_rejects_read_only(tmp_path: Path) -> None:
    workbook_path = tmp_path / "rename-ro.xlsx"
    _create_workbook(workbook_path)
    store = WorkbookStore()

    handle = workbook_open(store, path=str(workbook_path), read_only=True)

    try:
        sheet_rename(store, workbook_id=handle.workbook_id, sheet="Sheet1", new_name="New")
    except XlsxAgentError as exc:
        assert exc.code == "unsupported_operation"
    else:
        raise AssertionError("Expected rename on read-only workbook to fail.")


# --- formula_set ---


def test_formula_set_writes_formulas(tmp_path: Path) -> None:
    workbook_path = tmp_path / "formula.xlsx"
    _create_workbook(workbook_path)
    store = WorkbookStore()

    handle = workbook_open(store, path=str(workbook_path))
    result = formula_set(
        store,
        workbook_id=handle.workbook_id,
        sheet="Sheet1",
        formulas={"C1": "=SUM(A2:B2)", "C2": "=A2&B2"},
    )
    read_result = range_read(
        store,
        workbook_id=handle.workbook_id,
        sheet="Sheet1",
        range="C1:C2",
    )

    assert result["written_count"] == 2
    assert result["dirty"] is True
    assert read_result.data[0][0]["value"]["type"] == "formula"
    assert read_result.data[0][0]["value"]["formula"] == "=SUM(A2:B2)"
    assert read_result.data[1][0]["value"]["type"] == "formula"


def test_formula_set_rejects_invalid_formula(tmp_path: Path) -> None:
    workbook_path = tmp_path / "formula-invalid.xlsx"
    _create_workbook(workbook_path)
    store = WorkbookStore()

    handle = workbook_open(store, path=str(workbook_path))

    try:
        formula_set(
            store,
            workbook_id=handle.workbook_id,
            sheet="Sheet1",
            formulas={"A3": "no-equals-sign"},
        )
    except XlsxAgentError as exc:
        assert exc.code == "invalid_input"
    else:
        raise AssertionError("Expected invalid formula to fail.")


def test_formula_set_rejects_empty_formula(tmp_path: Path) -> None:
    workbook_path = tmp_path / "formula-empty.xlsx"
    _create_workbook(workbook_path)
    store = WorkbookStore()

    handle = workbook_open(store, path=str(workbook_path))

    try:
        formula_set(
            store,
            workbook_id=handle.workbook_id,
            sheet="Sheet1",
            formulas={"A3": "="},
        )
    except XlsxAgentError as exc:
        assert exc.code == "invalid_input"
    else:
        raise AssertionError("Expected empty formula to fail.")


def test_formula_set_rejects_read_only(tmp_path: Path) -> None:
    workbook_path = tmp_path / "formula-ro.xlsx"
    _create_workbook(workbook_path)
    store = WorkbookStore()

    handle = workbook_open(store, path=str(workbook_path), read_only=True)

    try:
        formula_set(
            store,
            workbook_id=handle.workbook_id,
            sheet="Sheet1",
            formulas={"C1": "=1+1"},
        )
    except XlsxAgentError as exc:
        assert exc.code == "unsupported_operation"
    else:
        raise AssertionError("Expected formula_set on read-only workbook to fail.")


# --- table_write ---


def test_table_write_list_rows(tmp_path: Path) -> None:
    workbook_path = tmp_path / "table-list.xlsx"
    _create_workbook(workbook_path)
    store = WorkbookStore()

    handle = workbook_open(store, path=str(workbook_path))
    result = table_write(
        store,
        workbook_id=handle.workbook_id,
        sheet="Sheet1",
        start_cell="A4",
        rows=[["Dave", 40], ["Eve", 50]],
    )
    read_result = range_read(
        store,
        workbook_id=handle.workbook_id,
        sheet="Sheet1",
        range="A4:B5",
        format="matrix",
    )

    assert result["rows_written"] == 2
    assert result["columns_written"] == 2
    assert result["dirty"] is True
    assert read_result.data == [["Dave", 40], ["Eve", 50]]


def test_table_write_dict_rows(tmp_path: Path) -> None:
    workbook_path = tmp_path / "table-dict.xlsx"
    _create_workbook(workbook_path)
    store = WorkbookStore()

    handle = workbook_open(store, path=str(workbook_path))
    result = table_write(
        store,
        workbook_id=handle.workbook_id,
        sheet="Sheet1",
        start_cell="A4",
        rows=[{"Name": "Dave", "Amount": 40}, {"Name": "Eve", "Amount": 50}],
    )
    read_result = range_read(
        store,
        workbook_id=handle.workbook_id,
        sheet="Sheet1",
        range="A4:B5",
        format="matrix",
    )

    assert result["rows_written"] == 2
    assert result["columns_written"] == 2
    assert read_result.data == [["Dave", 40], ["Eve", 50]]


def test_table_write_dict_rows_with_custom_columns(tmp_path: Path) -> None:
    workbook_path = tmp_path / "table-cols.xlsx"
    _create_workbook(workbook_path)
    store = WorkbookStore()

    handle = workbook_open(store, path=str(workbook_path))
    result = table_write(
        store,
        workbook_id=handle.workbook_id,
        sheet="Sheet1",
        start_cell="A4",
        rows=[{"Name": "Dave", "Amount": 40, "Extra": 99}],
        columns=["Amount", "Name"],
    )
    read_result = range_read(
        store,
        workbook_id=handle.workbook_id,
        sheet="Sheet1",
        range="A4:B4",
        format="matrix",
    )

    assert result["columns_written"] == 2
    assert read_result.data == [[40, "Dave"]]


def test_table_write_overwrite_error_on_occupied_cells(tmp_path: Path) -> None:
    workbook_path = tmp_path / "table-overwrite.xlsx"
    _create_workbook(workbook_path)
    store = WorkbookStore()

    handle = workbook_open(store, path=str(workbook_path))

    try:
        table_write(
            store,
            workbook_id=handle.workbook_id,
            sheet="Sheet1",
            start_cell="A1",
            rows=[["New", 999]],
            overwrite="error",
        )
    except XlsxAgentError as exc:
        assert exc.code == "target_not_empty"
    else:
        raise AssertionError("Expected overwrite error on occupied cells to fail.")


def test_table_write_overwrite_replace_on_occupied_cells(tmp_path: Path) -> None:
    workbook_path = tmp_path / "table-replace.xlsx"
    _create_workbook(workbook_path)
    store = WorkbookStore()

    handle = workbook_open(store, path=str(workbook_path))
    result = table_write(
        store,
        workbook_id=handle.workbook_id,
        sheet="Sheet1",
        start_cell="A1",
        rows=[["New", 999]],
        overwrite="replace",
    )
    read_result = range_read(
        store,
        workbook_id=handle.workbook_id,
        sheet="Sheet1",
        range="A1:B1",
        format="matrix",
    )

    assert result["rows_written"] == 1
    assert read_result.data == [["New", 999]]


def test_table_write_rejects_read_only(tmp_path: Path) -> None:
    workbook_path = tmp_path / "table-ro.xlsx"
    _create_workbook(workbook_path)
    store = WorkbookStore()

    handle = workbook_open(store, path=str(workbook_path), read_only=True)

    try:
        table_write(
            store,
            workbook_id=handle.workbook_id,
            sheet="Sheet1",
            start_cell="A4",
            rows=[["Dave", 40]],
        )
    except XlsxAgentError as exc:
        assert exc.code == "unsupported_operation"
    else:
        raise AssertionError("Expected table_write on read-only workbook to fail.")


# --- rows_modify ---


def test_rows_modify_overwrites_rows(tmp_path: Path) -> None:
    workbook_path = tmp_path / "rows-mod.xlsx"
    _create_workbook(workbook_path)
    store = WorkbookStore()

    handle = workbook_open(store, path=str(workbook_path))
    result = rows_modify(
        store,
        workbook_id=handle.workbook_id,
        sheet="Sheet1",
        start_row=2,
        rows=[["Bob", 20], ["Carol", 30]],
    )
    read_result = range_read(
        store,
        workbook_id=handle.workbook_id,
        sheet="Sheet1",
        range="A2:B3",
        format="matrix",
    )

    assert result["modified_count"] == 2
    assert result["dirty"] is True
    assert read_result.data == [["Bob", 20], ["Carol", 30]]


def test_rows_modify_dict_rows(tmp_path: Path) -> None:
    workbook_path = tmp_path / "rows-mod-dict.xlsx"
    _create_workbook(workbook_path)
    store = WorkbookStore()

    handle = workbook_open(store, path=str(workbook_path))
    result = rows_modify(
        store,
        workbook_id=handle.workbook_id,
        sheet="Sheet1",
        start_row=2,
        rows=[{"Name": "Bob", "Amount": 20}],
    )
    read_result = range_read(
        store,
        workbook_id=handle.workbook_id,
        sheet="Sheet1",
        range="A2:B2",
        format="matrix",
    )

    assert result["modified_count"] == 1
    assert read_result.data == [["Bob", 20]]


def test_rows_modify_rejects_invalid_start_row(tmp_path: Path) -> None:
    workbook_path = tmp_path / "rows-mod-bad.xlsx"
    _create_workbook(workbook_path)
    store = WorkbookStore()

    handle = workbook_open(store, path=str(workbook_path))

    try:
        rows_modify(
            store,
            workbook_id=handle.workbook_id,
            sheet="Sheet1",
            start_row=0,
            rows=[["Bob", 20]],
        )
    except XlsxAgentError as exc:
        assert exc.code == "invalid_input"
    else:
        raise AssertionError("Expected invalid start_row to fail.")


def test_rows_modify_rejects_read_only(tmp_path: Path) -> None:
    workbook_path = tmp_path / "rows-mod-ro.xlsx"
    _create_workbook(workbook_path)
    store = WorkbookStore()

    handle = workbook_open(store, path=str(workbook_path), read_only=True)

    try:
        rows_modify(
            store,
            workbook_id=handle.workbook_id,
            sheet="Sheet1",
            start_row=2,
            rows=[["Bob", 20]],
        )
    except XlsxAgentError as exc:
        assert exc.code == "unsupported_operation"
    else:
        raise AssertionError("Expected rows_modify on read-only workbook to fail.")


# --- columns_modify ---


def test_columns_modify_overwrites_columns_by_int(tmp_path: Path) -> None:
    workbook_path = tmp_path / "cols-mod-int.xlsx"
    _create_workbook(workbook_path)
    store = WorkbookStore()

    handle = workbook_open(store, path=str(workbook_path))
    result = columns_modify(
        store,
        workbook_id=handle.workbook_id,
        sheet="Sheet1",
        start_column=1,
        columns=[["X", "Y", "Z"]],
    )
    read_result = range_read(
        store,
        workbook_id=handle.workbook_id,
        sheet="Sheet1",
        range="A1:A3",
        format="matrix",
    )

    assert result["modified_count"] == 1
    assert result["start_column"] == "A"
    assert result["dirty"] is True
    assert read_result.data == [["X"], ["Y"], ["Z"]]


def test_columns_modify_overwrites_columns_by_letter(tmp_path: Path) -> None:
    workbook_path = tmp_path / "cols-mod-letter.xlsx"
    _create_workbook(workbook_path)
    store = WorkbookStore()

    handle = workbook_open(store, path=str(workbook_path))
    result = columns_modify(
        store,
        workbook_id=handle.workbook_id,
        sheet="Sheet1",
        start_column="B",
        columns=[["P", "Q"]],
    )
    read_result = range_read(
        store,
        workbook_id=handle.workbook_id,
        sheet="Sheet1",
        range="B1:B2",
        format="matrix",
    )

    assert result["modified_count"] == 1
    assert result["start_column"] == "B"
    assert read_result.data == [["P"], ["Q"]]


def test_columns_modify_dict_columns(tmp_path: Path) -> None:
    workbook_path = tmp_path / "cols-mod-dict.xlsx"
    _create_workbook(workbook_path)
    store = WorkbookStore()

    handle = workbook_open(store, path=str(workbook_path))
    result = columns_modify(
        store,
        workbook_id=handle.workbook_id,
        sheet="Sheet1",
        start_column=1,
        columns=[{"Row1": "Alice", "Row2": 10}],
    )
    read_result = range_read(
        store,
        workbook_id=handle.workbook_id,
        sheet="Sheet1",
        range="A1:A2",
        format="matrix",
    )

    assert result["modified_count"] == 1
    assert read_result.data == [["Alice"], [10]]


def test_columns_modify_rejects_invalid_column(tmp_path: Path) -> None:
    workbook_path = tmp_path / "cols-mod-bad.xlsx"
    _create_workbook(workbook_path)
    store = WorkbookStore()

    handle = workbook_open(store, path=str(workbook_path))

    try:
        columns_modify(
            store,
            workbook_id=handle.workbook_id,
            sheet="Sheet1",
            start_column=0,
            columns=[["X"]],
        )
    except XlsxAgentError as exc:
        assert exc.code == "invalid_input"
    else:
        raise AssertionError("Expected invalid column index to fail.")


def test_columns_modify_rejects_read_only(tmp_path: Path) -> None:
    workbook_path = tmp_path / "cols-mod-ro.xlsx"
    _create_workbook(workbook_path)
    store = WorkbookStore()

    handle = workbook_open(store, path=str(workbook_path), read_only=True)

    try:
        columns_modify(
            store,
            workbook_id=handle.workbook_id,
            sheet="Sheet1",
            start_column=1,
            columns=[["X"]],
        )
    except XlsxAgentError as exc:
        assert exc.code == "unsupported_operation"
    else:
        raise AssertionError("Expected columns_modify on read-only workbook to fail.")
