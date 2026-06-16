from __future__ import annotations

from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import coordinate_to_tuple

from xlsx_agent.core.errors import XlsxAgentError
from xlsx_agent.core.sheets import get_worksheet
from xlsx_agent.core.store import WorkbookStore


def _set_literal_value(cell, value: object) -> None:
    cell.value = value
    if isinstance(value, str) and value.startswith("="):
        cell.data_type = "s"


def cells_write(
    store: WorkbookStore,
    *,
    workbook_id: str,
    sheet: str,
    cells: dict[str, object],
) -> dict[str, object]:
    session = store.get(workbook_id)
    if session.read_only:
        raise XlsxAgentError(
            code="unsupported_operation",
            message=f"Workbook '{workbook_id}' was opened in read-only mode.",
        )

    worksheet = get_worksheet(session.workbook, sheet)
    normalized_items = []
    for address, value in cells.items():
        try:
            row, column = coordinate_to_tuple(address)
        except ValueError as exc:
            raise XlsxAgentError(
                code="invalid_input",
                message=f"Cell address '{address}' is invalid.",
            ) from exc
        if row < 1 or column < 1:
            raise XlsxAgentError(
                code="invalid_input",
                message=f"Cell address '{address}' is invalid.",
            )
        normalized_items.append((address.upper(), value))

    for address, value in normalized_items:
        _set_literal_value(worksheet[address], value)

    session.dirty = True
    return {
        "workbook_id": workbook_id,
        "sheet": sheet,
        "written_count": len(normalized_items),
        "dirty": session.dirty,
    }


def rows_append(
    store: WorkbookStore,
    *,
    workbook_id: str,
    sheet: str,
    rows: list[list[object] | dict[str, object]],
) -> dict[str, object]:
    session = store.get(workbook_id)
    if session.read_only:
        raise XlsxAgentError(
            code="unsupported_operation",
            message=f"Workbook '{workbook_id}' was opened in read-only mode.",
        )

    worksheet = get_worksheet(session.workbook, sheet)
    if not rows:
        return {
            "workbook_id": workbook_id,
            "sheet": sheet,
            "appended_count": 0,
            "dirty": session.dirty,
        }

    first_row = rows[0]
    if isinstance(first_row, dict):
        columns = list(first_row.keys())
        for row in rows:
            if not isinstance(row, dict):
                raise XlsxAgentError(
                    code="invalid_input",
                    message="All appended rows must use the same shape.",
                )
            worksheet.append([row.get(column) for column in columns])
    else:
        for row in rows:
            if isinstance(row, dict):
                raise XlsxAgentError(
                    code="invalid_input",
                    message="All appended rows must use the same shape.",
                )
            worksheet.append(row)

    session.dirty = True
    return {
        "workbook_id": workbook_id,
        "sheet": sheet,
        "appended_count": len(rows),
        "dirty": session.dirty,
    }


def _resolve_column_index(col: int | str) -> int:
    if isinstance(col, int):
        if col < 1:
            raise XlsxAgentError(
                code="invalid_input",
                message="Column index must be greater than or equal to 1.",
            )
        return col
    from openpyxl.utils import column_index_from_string

    try:
        return column_index_from_string(col)
    except ValueError as exc:
        raise XlsxAgentError(
            code="invalid_input",
            message=f"Column reference '{col}' is invalid.",
        ) from exc


def _resolve_start_cell(start_cell: str) -> tuple[int, int]:
    try:
        row, column = coordinate_to_tuple(start_cell)
    except ValueError as exc:
        raise XlsxAgentError(
            code="invalid_input",
            message=f"Cell address '{start_cell}' is invalid.",
        ) from exc
    if row < 1 or column < 1:
        raise XlsxAgentError(
            code="invalid_input",
            message=f"Cell address '{start_cell}' is invalid.",
        )
    return row, column


def table_write(
    store: WorkbookStore,
    *,
    workbook_id: str,
    sheet: str,
    start_cell: str,
    rows: list[list[object] | dict[str, object]],
    columns: list[str] | None = None,
    overwrite: str = "error",
) -> dict[str, object]:
    session = store.get(workbook_id)
    if session.read_only:
        raise XlsxAgentError(
            code="unsupported_operation",
            message=f"Workbook '{workbook_id}' was opened in read-only mode.",
        )

    worksheet = get_worksheet(session.workbook, sheet)

    if not rows:
        return {
            "workbook_id": workbook_id,
            "sheet": sheet,
            "start_cell": start_cell.upper(),
            "rows_written": 0,
            "columns_written": 0,
            "dirty": session.dirty,
        }

    start_row, start_col = _resolve_start_cell(start_cell)

    first_row = rows[0]
    if isinstance(first_row, dict):
        col_keys = columns if columns is not None else list(first_row.keys())
        num_cols = len(col_keys)
        for r_idx, row in enumerate(rows):
            if not isinstance(row, dict):
                raise XlsxAgentError(
                    code="invalid_input",
                    message="All rows must use the same shape.",
                )
            for c_idx, key in enumerate(col_keys):
                cell = worksheet.cell(row=start_row + r_idx, column=start_col + c_idx)
                if overwrite == "error" and cell.value is not None:
                    raise XlsxAgentError(
                        code="target_not_empty",
                        message=f"Cell at row {start_row + r_idx}, column {start_col + c_idx} is not empty.",
                    )
                cell.value = row.get(key)
    else:
        num_cols = max(len(row) for row in rows) if rows else 0
        for r_idx, row in enumerate(rows):
            if isinstance(row, dict):
                raise XlsxAgentError(
                    code="invalid_input",
                    message="All rows must use the same shape.",
                )
            for c_idx, value in enumerate(row):
                cell = worksheet.cell(row=start_row + r_idx, column=start_col + c_idx)
                if overwrite == "error" and cell.value is not None:
                    raise XlsxAgentError(
                        code="target_not_empty",
                        message=f"Cell at row {start_row + r_idx}, column {start_col + c_idx} is not empty.",
                    )
                cell.value = value

    session.dirty = True
    return {
        "workbook_id": workbook_id,
        "sheet": sheet,
        "start_cell": start_cell.upper(),
        "rows_written": len(rows),
        "columns_written": num_cols,
        "dirty": session.dirty,
    }


def rows_modify(
    store: WorkbookStore,
    *,
    workbook_id: str,
    sheet: str,
    start_row: int,
    rows: list[list[object] | dict[str, object]],
) -> dict[str, object]:
    session = store.get(workbook_id)
    if session.read_only:
        raise XlsxAgentError(
            code="unsupported_operation",
            message=f"Workbook '{workbook_id}' was opened in read-only mode.",
        )

    worksheet = get_worksheet(session.workbook, sheet)

    if start_row < 1:
        raise XlsxAgentError(
            code="invalid_input",
            message="start_row must be greater than or equal to 1.",
        )

    if not rows:
        return {
            "workbook_id": workbook_id,
            "sheet": sheet,
            "start_row": start_row,
            "modified_count": 0,
            "dirty": session.dirty,
        }

    first_row = rows[0]
    if isinstance(first_row, dict):
        columns = list(first_row.keys())
        for r_idx, row in enumerate(rows):
            if not isinstance(row, dict):
                raise XlsxAgentError(
                    code="invalid_input",
                    message="All rows must use the same shape.",
                )
            for c_idx, key in enumerate(columns):
                worksheet.cell(row=start_row + r_idx, column=c_idx + 1).value = row.get(key)
    else:
        for r_idx, row in enumerate(rows):
            if isinstance(row, dict):
                raise XlsxAgentError(
                    code="invalid_input",
                    message="All rows must use the same shape.",
                )
            for c_idx, value in enumerate(row):
                worksheet.cell(row=start_row + r_idx, column=c_idx + 1).value = value

    session.dirty = True
    return {
        "workbook_id": workbook_id,
        "sheet": sheet,
        "start_row": start_row,
        "modified_count": len(rows),
        "dirty": session.dirty,
    }


def columns_modify(
    store: WorkbookStore,
    *,
    workbook_id: str,
    sheet: str,
    start_column: int | str,
    columns: list[list[object] | dict[str, object]],
) -> dict[str, object]:
    session = store.get(workbook_id)
    if session.read_only:
        raise XlsxAgentError(
            code="unsupported_operation",
            message=f"Workbook '{workbook_id}' was opened in read-only mode.",
        )

    worksheet = get_worksheet(session.workbook, sheet)

    col_index = _resolve_column_index(start_column)

    if not columns:
        return {
            "workbook_id": workbook_id,
            "sheet": sheet,
            "start_column": start_column if isinstance(start_column, str) else get_column_letter(start_column),
            "modified_count": 0,
            "dirty": session.dirty,
        }

    first_col = columns[0]
    if isinstance(first_col, dict):
        row_keys = list(first_col.keys())
        for c_idx, col_data in enumerate(columns):
            if not isinstance(col_data, dict):
                raise XlsxAgentError(
                    code="invalid_input",
                    message="All columns must use the same shape.",
                )
            for r_idx, key in enumerate(row_keys):
                worksheet.cell(row=r_idx + 1, column=col_index + c_idx).value = col_data.get(key)
    else:
        for c_idx, col_data in enumerate(columns):
            if isinstance(col_data, dict):
                raise XlsxAgentError(
                    code="invalid_input",
                    message="All columns must use the same shape.",
                )
            for r_idx, value in enumerate(col_data):
                worksheet.cell(row=r_idx + 1, column=col_index + c_idx).value = value

    session.dirty = True
    return {
        "workbook_id": workbook_id,
        "sheet": sheet,
        "start_column": start_column if isinstance(start_column, str) else get_column_letter(start_column),
        "modified_count": len(columns),
        "dirty": session.dirty,
    }
