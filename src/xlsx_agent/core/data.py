from __future__ import annotations

from datetime import date, datetime

from openpyxl.utils import get_column_letter, range_boundaries

from xlsx_agent.core.errors import XlsxAgentError
from xlsx_agent.core.sheets import get_worksheet
from xlsx_agent.core.store import WorkbookStore


def _resolve_range(range_string: str) -> tuple[int, int, int, int, str]:
    try:
        min_col, min_row, max_col, max_row = range_boundaries(range_string)
    except ValueError as exc:
        raise XlsxAgentError(
            code="invalid_range",
            message=f"Range '{range_string}' is invalid.",
        ) from exc
    resolved_range = (
        f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}"
    )
    return min_col, min_row, max_col, max_row, resolved_range


def _cell_to_value(cell) -> object:
    v = cell.value
    if v is None:
        return None
    if isinstance(v, bool):
        return v
    if isinstance(v, (int, float)):
        return v
    if isinstance(v, (date, datetime)):
        return v.isoformat()
    return str(v)


def range_find(
    store: WorkbookStore,
    *,
    workbook_id: str,
    sheet: str,
    range: str,
    value: object,
    match_type: str = "exact",
) -> dict[str, object]:
    session = store.get(workbook_id)
    worksheet = get_worksheet(session.workbook, sheet)
    min_col, min_row, max_col, max_row, resolved = _resolve_range(range)

    matches = []
    for row in worksheet.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            cell_val = _cell_to_value(cell)
            if match_type == "exact" and cell_val == value:
                matches.append(cell.coordinate)
            elif match_type == "contains" and cell_val is not None and str(value) in str(cell_val):
                matches.append(cell.coordinate)

    return {
        "workbook_id": workbook_id,
        "sheet": sheet,
        "range": resolved,
        "matches": matches,
        "match_count": len(matches),
    }


def range_replace(
    store: WorkbookStore,
    *,
    workbook_id: str,
    sheet: str,
    range: str,
    find: str,
    replace_with: str,
) -> dict[str, object]:
    session = store.get(workbook_id)
    if session.read_only:
        raise XlsxAgentError(
            code="unsupported_operation",
            message=f"Workbook '{workbook_id}' was opened in read-only mode.",
        )

    worksheet = get_worksheet(session.workbook, sheet)
    min_col, min_row, max_col, max_row, resolved = _resolve_range(range)

    replaced = 0
    for row in worksheet.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            if cell.value is not None and isinstance(cell.value, str) and find in cell.value:
                cell.value = cell.value.replace(find, replace_with)
                replaced += 1

    session.dirty = True
    return {
        "workbook_id": workbook_id,
        "sheet": sheet,
        "range": resolved,
        "replaced_count": replaced,
        "dirty": session.dirty,
    }


def distinct_values(
    store: WorkbookStore,
    *,
    workbook_id: str,
    sheet: str,
    range: str,
) -> dict[str, object]:
    session = store.get(workbook_id)
    worksheet = get_worksheet(session.workbook, sheet)
    min_col, min_row, max_col, max_row, resolved = _resolve_range(range)

    values = set()
    for row in worksheet.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            v = _cell_to_value(cell)
            if v is not None:
                values.add(v)

    return {
        "workbook_id": workbook_id,
        "sheet": sheet,
        "range": resolved,
        "values": sorted(values, key=str),
        "count": len(values),
    }


def duplicates_find(
    store: WorkbookStore,
    *,
    workbook_id: str,
    sheet: str,
    range: str,
) -> dict[str, object]:
    session = store.get(workbook_id)
    worksheet = get_worksheet(session.workbook, sheet)
    min_col, min_row, max_col, max_row, resolved = _resolve_range(range)

    from collections import Counter
    values = []
    coords = []
    for row in worksheet.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            v = _cell_to_value(cell)
            if v is not None:
                values.append(v)
                coords.append(cell.coordinate)

    counts = Counter(values)
    duplicates = {v: [coords[i] for i, val in enumerate(values) if val == v] for v, c in counts.items() if c > 1}

    return {
        "workbook_id": workbook_id,
        "sheet": sheet,
        "range": resolved,
        "duplicates": duplicates,
        "duplicate_value_count": len(duplicates),
    }


def blanks_find(
    store: WorkbookStore,
    *,
    workbook_id: str,
    sheet: str,
    range: str,
) -> dict[str, object]:
    session = store.get(workbook_id)
    worksheet = get_worksheet(session.workbook, sheet)
    min_col, min_row, max_col, max_row, resolved = _resolve_range(range)

    blanks = []
    for row in worksheet.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            if cell.value is None:
                blanks.append(cell.coordinate)

    return {
        "workbook_id": workbook_id,
        "sheet": sheet,
        "range": resolved,
        "blank_cells": blanks,
        "blank_count": len(blanks),
    }


def errors_find(
    store: WorkbookStore,
    *,
    workbook_id: str,
    sheet: str,
    range: str,
) -> dict[str, object]:
    session = store.get(workbook_id)
    worksheet = get_worksheet(session.workbook, sheet)
    min_col, min_row, max_col, max_row, resolved = _resolve_range(range)

    errors = []
    for row in worksheet.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            if cell.data_type == "e":
                errors.append({"cell": cell.coordinate, "error": str(cell.value)})

    return {
        "workbook_id": workbook_id,
        "sheet": sheet,
        "range": resolved,
        "error_cells": errors,
        "error_count": len(errors),
    }


def range_sort(
    store: WorkbookStore,
    *,
    workbook_id: str,
    sheet: str,
    range: str,
    sort_column: int | str,
    ascending: bool = True,
) -> dict[str, object]:
    session = store.get(workbook_id)
    if session.read_only:
        raise XlsxAgentError(
            code="unsupported_operation",
            message=f"Workbook '{workbook_id}' was opened in read-only mode.",
        )

    worksheet = get_worksheet(session.workbook, sheet)
    min_col, min_row, max_col, max_row, resolved = _resolve_range(range)

    if isinstance(sort_column, str):
        from openpyxl.utils import column_index_from_string
        sort_col_idx = column_index_from_string(sort_column)
    else:
        sort_col_idx = sort_column

    data = []
    for row in worksheet.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        data.append([cell.value for cell in row])

    sort_key_idx = sort_col_idx - min_col
    if sort_key_idx < 0 or sort_key_idx >= len(data[0]) if data else 0:
        raise XlsxAgentError(
            code="invalid_input",
            message="sort_column is outside the range.",
        )

    data.sort(key=lambda r: (r[sort_key_idx] is None, r[sort_key_idx] if r[sort_key_idx] is not None else ""), reverse=not ascending)

    for r_idx, row_data in enumerate(data):
        for c_idx, value in enumerate(row_data):
            worksheet.cell(row=min_row + r_idx, column=min_col + c_idx).value = value

    session.dirty = True
    return {
        "workbook_id": workbook_id,
        "sheet": sheet,
        "range": resolved,
        "sorted_rows": len(data),
        "dirty": session.dirty,
    }


def range_deduplicate(
    store: WorkbookStore,
    *,
    workbook_id: str,
    sheet: str,
    range: str,
) -> dict[str, object]:
    session = store.get(workbook_id)
    if session.read_only:
        raise XlsxAgentError(
            code="unsupported_operation",
            message=f"Workbook '{workbook_id}' was opened in read-only mode.",
        )

    worksheet = get_worksheet(session.workbook, sheet)
    min_col, min_row, max_col, max_row, resolved = _resolve_range(range)

    data = []
    for row in worksheet.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        data.append([cell.value for cell in row])

    seen = set()
    unique_rows = []
    for row in data:
        key = tuple(str(v) for v in row)
        if key not in seen:
            seen.add(key)
            unique_rows.append(row)

    removed = len(data) - len(unique_rows)

    for r_idx, row_data in enumerate(unique_rows):
        for c_idx, value in enumerate(row_data):
            worksheet.cell(row=min_row + r_idx, column=min_col + c_idx).value = value

    for r_idx in range(len(unique_rows), len(data)):
        for c_idx in range(max_col - min_col + 1):
            worksheet.cell(row=min_row + r_idx, column=min_col + c_idx).value = None

    session.dirty = True
    return {
        "workbook_id": workbook_id,
        "sheet": sheet,
        "range": resolved,
        "removed_duplicates": removed,
        "remaining_rows": len(unique_rows),
        "dirty": session.dirty,
    }
