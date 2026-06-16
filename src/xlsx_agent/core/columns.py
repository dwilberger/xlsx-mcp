from __future__ import annotations

from openpyxl.utils import column_index_from_string, get_column_letter

from xlsx_agent.core.errors import XlsxAgentError
from xlsx_agent.core.sheets import get_worksheet
from xlsx_agent.core.store import WorkbookStore


def _resolve_column_index(col: int | str) -> int:
    if isinstance(col, int):
        if col < 1:
            raise XlsxAgentError(
                code="invalid_input",
                message="Column index must be greater than or equal to 1.",
            )
        return col
    try:
        return column_index_from_string(col)
    except ValueError as exc:
        raise XlsxAgentError(
            code="invalid_input",
            message=f"Column reference '{col}' is invalid.",
        ) from exc


def columns_insert(
    store: WorkbookStore,
    *,
    workbook_id: str,
    sheet: str,
    start_column: int | str,
    count: int = 1,
) -> dict[str, object]:
    session = store.get(workbook_id)
    if session.read_only:
        raise XlsxAgentError(
            code="unsupported_operation",
            message=f"Workbook '{workbook_id}' was opened in read-only mode.",
        )

    if count < 1:
        raise XlsxAgentError(
            code="invalid_input",
            message="count must be greater than or equal to 1.",
        )

    col_index = _resolve_column_index(start_column)
    worksheet = get_worksheet(session.workbook, sheet)
    worksheet.insert_cols(col_index, count)

    session.dirty = True
    return {
        "workbook_id": workbook_id,
        "sheet": sheet,
        "start_column": get_column_letter(col_index),
        "inserted_count": count,
        "dirty": session.dirty,
    }


def columns_delete(
    store: WorkbookStore,
    *,
    workbook_id: str,
    sheet: str,
    start_column: int | str,
    count: int = 1,
) -> dict[str, object]:
    session = store.get(workbook_id)
    if session.read_only:
        raise XlsxAgentError(
            code="unsupported_operation",
            message=f"Workbook '{workbook_id}' was opened in read-only mode.",
        )

    if count < 1:
        raise XlsxAgentError(
            code="invalid_input",
            message="count must be greater than or equal to 1.",
        )

    col_index = _resolve_column_index(start_column)
    worksheet = get_worksheet(session.workbook, sheet)
    worksheet.delete_cols(col_index, count)

    session.dirty = True
    return {
        "workbook_id": workbook_id,
        "sheet": sheet,
        "start_column": get_column_letter(col_index),
        "deleted_count": count,
        "dirty": session.dirty,
    }
