from __future__ import annotations

from openpyxl.utils import column_index_from_string, get_column_letter

from xlsx_agent.core.errors import XlsxAgentError
from xlsx_agent.core.sheets import get_worksheet
from xlsx_agent.core.store import WorkbookStore


def formula_read(
    store: WorkbookStore,
    *,
    workbook_id: str,
    sheet: str,
    cell: str,
) -> dict[str, object]:
    session = store.get(workbook_id)
    worksheet = get_worksheet(session.workbook, sheet)

    c = worksheet[cell.upper()]
    if c.data_type != "f":
        return {
            "workbook_id": workbook_id,
            "sheet": sheet,
            "cell": cell.upper(),
            "has_formula": False,
            "formula": None,
            "cached_value": c.value,
        }

    return {
        "workbook_id": workbook_id,
        "sheet": sheet,
        "cell": cell.upper(),
        "has_formula": True,
        "formula": str(c.value),
        "cached_value": None,
    }


def formula_copy(
    store: WorkbookStore,
    *,
    workbook_id: str,
    sheet: str,
    source: str,
    target_range: str,
) -> dict[str, object]:
    session = store.get(workbook_id)
    if session.read_only:
        raise XlsxAgentError(
            code="unsupported_operation",
            message=f"Workbook '{workbook_id}' was opened in read-only mode.",
        )

    worksheet = get_worksheet(session.workbook, sheet)

    from openpyxl.utils import range_boundaries
    try:
        min_col, min_row, max_col, max_row, resolved = range_boundaries(target_range)
    except ValueError as exc:
        raise XlsxAgentError(
            code="invalid_range",
            message=f"Range '{target_range}' is invalid.",
        ) from exc

    source_cell = worksheet[source.upper()]
    if source_cell.data_type != "f":
        raise XlsxAgentError(
            code="invalid_input",
            message=f"Cell '{source}' does not contain a formula.",
        )

    formula = str(source_cell.value)
    copied = 0
    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            cell = worksheet.cell(row=row, column=col)
            if cell.coordinate != source.upper():
                cell.value = formula
                copied += 1

    session.dirty = True
    return {
        "workbook_id": workbook_id,
        "sheet": sheet,
        "source": source.upper(),
        "target_range": resolved,
        "copied_count": copied,
        "dirty": session.dirty,
    }


def formula_fill_down(
    store: WorkbookStore,
    *,
    workbook_id: str,
    sheet: str,
    start_cell: str,
    end_row: int,
) -> dict[str, object]:
    session = store.get(workbook_id)
    if session.read_only:
        raise XlsxAgentError(
            code="unsupported_operation",
            message=f"Workbook '{workbook_id}' was opened in read-only mode.",
        )

    worksheet = get_worksheet(session.workbook, sheet)

    from openpyxl.utils.cell import coordinate_to_tuple
    try:
        start_row, col = coordinate_to_tuple(start_cell)
    except ValueError as exc:
        raise XlsxAgentError(
            code="invalid_input",
            message=f"Cell address '{start_cell}' is invalid.",
        ) from exc

    source_cell = worksheet.cell(row=start_row, column=col)
    if source_cell.data_type != "f":
        raise XlsxAgentError(
            code="invalid_input",
            message=f"Cell '{start_cell}' does not contain a formula.",
        )

    formula = str(source_cell.value)
    filled = 0
    for row in range(start_row + 1, end_row + 1):
        cell = worksheet.cell(row=row, column=col)
        cell.value = formula
        filled += 1

    session.dirty = True
    return {
        "workbook_id": workbook_id,
        "sheet": sheet,
        "start_cell": start_cell.upper(),
        "end_row": end_row,
        "filled_count": filled,
        "dirty": session.dirty,
    }


def formula_fill_right(
    store: WorkbookStore,
    *,
    workbook_id: str,
    sheet: str,
    start_cell: str,
    end_column: int | str,
) -> dict[str, object]:
    session = store.get(workbook_id)
    if session.read_only:
        raise XlsxAgentError(
            code="unsupported_operation",
            message=f"Workbook '{workbook_id}' was opened in read-only mode.",
        )

    worksheet = get_worksheet(session.workbook, sheet)

    from openpyxl.utils.cell import coordinate_to_tuple
    try:
        row, start_col = coordinate_to_tuple(start_cell)
    except ValueError as exc:
        raise XlsxAgentError(
            code="invalid_input",
            message=f"Cell address '{start_cell}' is invalid.",
        ) from exc

    if isinstance(end_column, str):
        end_col = column_index_from_string(end_column)
    else:
        end_col = end_column

    source_cell = worksheet.cell(row=row, column=start_col)
    if source_cell.data_type != "f":
        raise XlsxAgentError(
            code="invalid_input",
            message=f"Cell '{start_cell}' does not contain a formula.",
        )

    formula = str(source_cell.value)
    filled = 0
    for col in range(start_col + 1, end_col + 1):
        cell = worksheet.cell(row=row, column=col)
        cell.value = formula
        filled += 1

    session.dirty = True
    return {
        "workbook_id": workbook_id,
        "sheet": sheet,
        "start_cell": start_cell.upper(),
        "end_column": get_column_letter(end_col),
        "filled_count": filled,
        "dirty": session.dirty,
    }


def formula_to_values(
    store: WorkbookStore,
    *,
    workbook_id: str,
    sheet: str,
    range: str,
) -> dict[str, object]:
    session = store.get(workbook_id)
    if session.read_only:
        raise XlsxAgentError(
            code="unsupported_operation",
            message=f"Workbook '{workbook_id}' was opened in read-only mode.",
        )

    worksheet = get_worksheet(session.workbook, sheet)

    from openpyxl.utils import range_boundaries
    try:
        min_col, min_row, max_col, max_row, resolved = range_boundaries(range)
    except ValueError as exc:
        raise XlsxAgentError(
            code="invalid_range",
            message=f"Range '{range}' is invalid.",
        ) from exc

    converted = 0
    for row in worksheet.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            if cell.data_type == "f":
                cell.value = cell.value
                cell.data_type = "n"
                converted += 1

    session.dirty = True
    return {
        "workbook_id": workbook_id,
        "sheet": sheet,
        "range": resolved,
        "converted_count": converted,
        "dirty": session.dirty,
    }


def array_formula_list(
    store: WorkbookStore,
    *,
    workbook_id: str,
    sheet: str,
) -> dict[str, object]:
    session = store.get(workbook_id)
    worksheet = get_worksheet(session.workbook, sheet)

    formulas = []
    for row in worksheet.iter_rows():
        for cell in row:
            if cell.data_type == "f" and cell.value and "{" in str(cell.value):
                formulas.append({
                    "cell": cell.coordinate,
                    "formula": str(cell.value),
                })

    return {
        "workbook_id": workbook_id,
        "sheet": sheet,
        "formulas": formulas,
        "count": len(formulas),
    }
