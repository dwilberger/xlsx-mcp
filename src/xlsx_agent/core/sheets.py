from __future__ import annotations

from openpyxl.utils import get_column_letter

from xlsx_agent.core.errors import XlsxAgentError
from xlsx_agent.core.store import WorkbookStore
from xlsx_agent.schemas.common import SheetMetadata


def _check_read_only(session) -> None:
    if session.read_only:
        raise XlsxAgentError(
            code="unsupported_operation",
            message=f"Workbook '{session.workbook_id}' was opened in read-only mode.",
        )


def get_worksheet(workbook, sheet_name: str):
    try:
        return workbook[sheet_name]
    except KeyError as exc:
        raise XlsxAgentError(
            code="sheet_not_found",
            message=f"Sheet '{sheet_name}' was not found.",
        ) from exc


def _normalize_visibility(sheet_state: str) -> str:
    if sheet_state == "veryHidden":
        return "very_hidden"
    return sheet_state


def _compute_used_range(worksheet) -> str | None:
    if worksheet.max_row == 1 and worksheet.max_column == 1 and worksheet["A1"].value is None:
        return None
    end_column = get_column_letter(worksheet.max_column)
    return f"A1:{end_column}{worksheet.max_row}"


def _sheet_metadata(workbook, worksheet, position: int) -> SheetMetadata:
    return SheetMetadata(
        name=worksheet.title,
        position=position,
        visibility=_normalize_visibility(worksheet.sheet_state),
        is_active=workbook.active is worksheet,
        used_range=_compute_used_range(worksheet),
        max_row=worksheet.max_row,
        max_column=worksheet.max_column,
    )


def sheet_list(store: WorkbookStore, *, workbook_id: str) -> list[SheetMetadata]:
    session = store.get(workbook_id)
    workbook = session.workbook
    return [
        _sheet_metadata(workbook, worksheet, position)
        for position, worksheet in enumerate(workbook.worksheets, start=1)
    ]


def sheet_create(
    store: WorkbookStore,
    *,
    workbook_id: str,
    name: str,
    position: int | None = None,
) -> SheetMetadata:
    session = store.get(workbook_id)
    _check_read_only(session)

    workbook = session.workbook
    if name in workbook.sheetnames:
        raise XlsxAgentError(
            code="invalid_input",
            message=f"Sheet '{name}' already exists.",
        )

    if position is not None and position < 1:
        raise XlsxAgentError(
            code="invalid_input",
            message="Sheet position must be greater than or equal to 1.",
        )

    index = None if position is None else position - 1
    worksheet = workbook.create_sheet(title=name, index=index)
    session.dirty = True
    actual_position = workbook.worksheets.index(worksheet) + 1
    return _sheet_metadata(workbook, worksheet, actual_position)


def sheet_delete(
    store: WorkbookStore,
    *,
    workbook_id: str,
    sheet: str,
) -> dict[str, object]:
    session = store.get(workbook_id)
    _check_read_only(session)

    worksheet = get_worksheet(session.workbook, sheet)

    if len(session.workbook.worksheets) <= 1:
        raise XlsxAgentError(
            code="invalid_input",
            message="Cannot delete the last sheet in a workbook.",
        )

    session.workbook.remove(worksheet)
    session.dirty = True
    return {
        "workbook_id": workbook_id,
        "deleted_sheet": sheet,
        "dirty": session.dirty,
    }


def sheet_rename(
    store: WorkbookStore,
    *,
    workbook_id: str,
    sheet: str,
    new_name: str,
) -> dict[str, object]:
    session = store.get(workbook_id)
    _check_read_only(session)

    worksheet = get_worksheet(session.workbook, sheet)

    if new_name in session.workbook.sheetnames and new_name != sheet:
        raise XlsxAgentError(
            code="invalid_input",
            message=f"Sheet '{new_name}' already exists.",
        )

    old_name = worksheet.title
    worksheet.title = new_name
    session.dirty = True
    return {
        "workbook_id": workbook_id,
        "old_name": old_name,
        "new_name": new_name,
        "dirty": session.dirty,
    }
