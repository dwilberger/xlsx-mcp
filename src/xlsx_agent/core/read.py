from __future__ import annotations

from datetime import date, datetime

from openpyxl.utils import get_column_letter, range_boundaries

from xlsx_agent.core.errors import XlsxAgentError
from xlsx_agent.core.sheets import get_worksheet
from xlsx_agent.core.store import WorkbookStore
from xlsx_agent.schemas.common import CellRead, CellValue, RangeReadResult


def _resolve_range(range_string: str) -> tuple[int, int, int, int, str]:
    try:
        min_col, min_row, max_col, max_row = range_boundaries(range_string)
    except ValueError as exc:
        raise XlsxAgentError(
            code="invalid_range",
            message=f"Range '{range_string}' is invalid.",
        ) from exc

    resolved_range = (
        f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}"
    )
    return min_col, min_row, max_col, max_row, resolved_range


def _serialize_cell_value(cell) -> CellValue:
    value = cell.value
    if value is None:
        return CellValue(type="empty")
    if cell.data_type == "f":
        return CellValue(type="formula", value=None, formula=str(value))
    if cell.data_type == "e":
        return CellValue(type="error", value=str(value))
    if isinstance(value, bool):
        return CellValue(type="boolean", value=value)
    if isinstance(value, datetime):
        if value.time() == datetime.min.time():
            return CellValue(type="date", value=value.date().isoformat())
        return CellValue(type="datetime", value=value.isoformat())
    if isinstance(value, date):
        return CellValue(type="date", value=value.isoformat())
    if isinstance(value, (int, float)):
        return CellValue(type="number", value=value)
    return CellValue(type="string", value=str(value))


def _project_primitive(cell_value: CellValue):
    if cell_value.type == "formula":
        return cell_value.formula
    if cell_value.type == "empty":
        return None
    return cell_value.value


def range_read(
    store: WorkbookStore,
    *,
    workbook_id: str,
    sheet: str,
    range: str,
    format: str = "cells",
) -> RangeReadResult:
    session = store.get(workbook_id)
    worksheet = get_worksheet(session.workbook, sheet)
    min_col, min_row, max_col, max_row, resolved_range = _resolve_range(range)

    rows = []
    for row in worksheet.iter_rows(
        min_row=min_row,
        max_row=max_row,
        min_col=min_col,
        max_col=max_col,
    ):
        rows.append(list(row))

    row_count = max_row - min_row + 1
    column_count = max_col - min_col + 1

    if format == "cells":
        data = []
        for row in rows:
            result_row = []
            for cell in row:
                result_row.append(
                    CellRead(address=cell.coordinate, value=_serialize_cell_value(cell)).model_dump()
                )
            data.append(result_row)
    elif format == "matrix":
        data = []
        for row in rows:
            data.append([_project_primitive(_serialize_cell_value(cell)) for cell in row])
    elif format == "records":
        header_cells = rows[0] if rows else []
        headers = [_project_primitive(_serialize_cell_value(cell)) for cell in header_cells]
        if any(not isinstance(header, str) or header == "" for header in headers):
            raise XlsxAgentError(
                code="invalid_input",
                message="Record headers must be non-empty strings.",
            )
        if len(set(headers)) != len(headers):
            raise XlsxAgentError(
                code="invalid_input",
                message="Record headers must be unique.",
            )

        data = []
        for row in rows[1:]:
            values = [_project_primitive(_serialize_cell_value(cell)) for cell in row]
            data.append(dict(zip(headers, values, strict=False)))
    else:
        raise XlsxAgentError(
            code="invalid_input",
            message=f"Range read format '{format}' is not supported.",
        )

    return RangeReadResult(
        sheet=sheet,
        requested_range=range,
        resolved_range=resolved_range,
        row_count=row_count,
        column_count=column_count,
        data=data,
    )
