from __future__ import annotations

from openpyxl.utils import get_column_letter, range_boundaries

from xlsx_agent.core.errors import XlsxAgentError
from xlsx_agent.core.sheets import get_worksheet
from xlsx_agent.core.store import WorkbookStore


def _resolve_range(range_string: str) -> tuple[int, int, int, int, str]:
    try:
        min_col, min_row, max_col, max_row = range_boundaries(range_string)
    except ValueError as exc:
        raise XlsxAgentError(
            code="invalid_range",
            message=f"Range '{range_string}' is invalid.",
        ) from exc
    resolved_range = (
        f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}"
    )
    return min_col, min_row, max_col, max_row, resolved_range


def range_write(
    store: WorkbookStore,
    *,
    workbook_id: str,
    sheet: str,
    range: str,
    data: list[list[object]],
) -> dict[str, object]:
    session = store.get(workbook_id)
    if session.read_only:
        raise XlsxAgentError(
            code="unsupported_operation",
            message=f"Workbook '{workbook_id}' was opened in read-only mode.",
        )

    worksheet = get_worksheet(session.workbook, sheet)
    min_col, min_row, max_col, max_row, resolved = _resolve_range(range)

    rows_written = 0
    cols_written = 0
    for r_idx, row in enumerate(data):
        for c_idx, value in enumerate(row):
            cell = worksheet.cell(row=min_row + r_idx, column=min_col + c_idx)
            cell.value = value
            if isinstance(value, str) and value.startswith("="):
                cell.data_type = "s"
        rows_written += 1
        cols_written = max(cols_written, len(row))

    session.dirty = True
    return {
        "workbook_id": workbook_id,
        "sheet": sheet,
        "range": resolved,
        "rows_written": rows_written,
        "columns_written": cols_written,
        "dirty": session.dirty,
    }


def range_clear(
    store: WorkbookStore,
    *,
    workbook_id: str,
    sheet: str,
    range: str,
) -> dict[str, object]:
    session = store.get(workbook_id)
    if session.read_only:
        raise XlsxAgentError(
            code="unsupported_operation",
            message=f"Workbook '{workbook_id}' was opened in read-only mode.",
        )

    worksheet = get_worksheet(session.workbook, sheet)
    min_col, min_row, max_col, max_row, resolved = _resolve_range(range)

    cleared = 0
    for row in worksheet.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            cell.value = None
            cell.data_type = "n"
            cleared += 1

    session.dirty = True
    return {
        "workbook_id": workbook_id,
        "sheet": sheet,
        "range": resolved,
        "cleared_count": cleared,
        "dirty": session.dirty,
    }
