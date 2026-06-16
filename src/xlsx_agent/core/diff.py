from __future__ import annotations

from xlsx_agent.core.errors import XlsxAgentError
from xlsx_agent.core.sheets import get_worksheet
from xlsx_agent.core.store import WorkbookStore


def _cell_to_value(cell) -> object:
    v = cell.value
    if v is None:
        return None
    if isinstance(v, bool):
        return v
    if isinstance(v, (int, float)):
        return v
    return str(v)


def sheet_diff(
    store: WorkbookStore,
    *,
    workbook_id: str,
    sheet_a: str,
    sheet_b: str,
) -> dict[str, object]:
    session = store.get(workbook_id)
    ws_a = get_worksheet(session.workbook, sheet_a)
    ws_b = get_worksheet(session.workbook, sheet_b)

    max_row = max(ws_a.max_row, ws_b.max_row)
    max_col = max(ws_a.max_column, ws_b.max_column)

    differences = []
    for row in range(1, max_row + 1):
        for col in range(1, max_col + 1):
            val_a = _cell_to_value(ws_a.cell(row=row, column=col))
            val_b = _cell_to_value(ws_b.cell(row=row, column=col))
            if val_a != val_b:
                differences.append({
                    "row": row,
                    "column": col,
                    "value_a": val_a,
                    "value_b": val_b,
                })

    return {
        "workbook_id": workbook_id,
        "sheet_a": sheet_a,
        "sheet_b": sheet_b,
        "differences": differences,
        "difference_count": len(differences),
    }


def workbook_diff(
    store: WorkbookStore,
    *,
    workbook_id: str,
    path_b: str,
) -> dict[str, object]:
    from openpyxl import load_workbook

    session = store.get(workbook_id)
    wb_a = session.workbook

    try:
        wb_b = load_workbook(path_b, read_only=True, data_only=True)
    except FileNotFoundError as exc:
        raise XlsxAgentError(
            code="file_not_found",
            message=f"File '{path_b}' was not found.",
        ) from exc

    sheets_a = set(wb_a.sheetnames)
    sheets_b = set(wb_b.sheetnames)

    result = {
        "workbook_id": workbook_id,
        "path_b": path_b,
        "only_in_a": sorted(sheets_a - sheets_b),
        "only_in_b": sorted(sheets_b - sheets_a),
        "common_sheets": [],
    }

    for sheet_name in sorted(sheets_a & sheets_b):
        ws_a = wb_a[sheet_name]
        ws_b = wb_b[sheet_name]

        max_row = max(ws_a.max_row, ws_b.max_row)
        max_col = max(ws_a.max_column, ws_b.max_column)

        diffs = []
        for row in range(1, max_row + 1):
            for col in range(1, max_col + 1):
                val_a = _cell_to_value(ws_a.cell(row=row, column=col))
                val_b = _cell_to_value(ws_b.cell(row=row, column=col))
                if val_a != val_b:
                    diffs.append({"row": row, "column": col, "value_a": val_a, "value_b": val_b})

        result["common_sheets"].append({
            "sheet": sheet_name,
            "difference_count": len(diffs),
            "differences": diffs[:100],
        })

    wb_b.close()
    return result
