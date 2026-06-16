from __future__ import annotations

from openpyxl.utils import get_column_letter

from xlsx_agent.core.errors import XlsxAgentError
from xlsx_agent.core.sheets import get_worksheet, _sheet_metadata
from xlsx_agent.core.store import WorkbookStore


def sheet_copy(
    store: WorkbookStore,
    *,
    workbook_id: str,
    sheet: str,
    new_name: str | None = None,
    position: int | None = None,
) -> dict[str, object]:
    session = store.get(workbook_id)
    if session.read_only:
        raise XlsxAgentError(
            code="unsupported_operation",
            message=f"Workbook '{workbook_id}' was opened in read-only mode.",
        )

    worksheet = get_worksheet(session.workbook, sheet)

    if new_name is None:
        new_name = f"{sheet} Copy"
    if new_name in session.workbook.sheetnames:
        raise XlsxAgentError(
            code="invalid_input",
            message=f"Sheet '{new_name}' already exists.",
        )

    index = None if position is None else position - 1
    new_ws = session.workbook.copy_worksheet(worksheet)
    new_ws.title = new_name
    if index is not None:
        session.workbook.move_sheet(new_ws, offset=index - session.workbook.worksheets.index(new_ws))

    session.dirty = True
    return {
        "workbook_id": workbook_id,
        "source_sheet": sheet,
        "new_sheet": new_name,
        "dirty": session.dirty,
    }


def sheet_used_range_get(
    store: WorkbookStore,
    *,
    workbook_id: str,
    sheet: str,
) -> dict[str, object]:
    session = store.get(workbook_id)
    worksheet = get_worksheet(session.workbook, sheet)

    if worksheet.max_row == 1 and worksheet.max_column == 1 and worksheet["A1"].value is None:
        used_range = None
    else:
        end_column = get_column_letter(worksheet.max_column)
        used_range = f"A1:{end_column}{worksheet.max_row}"

    return {
        "workbook_id": workbook_id,
        "sheet": sheet,
        "used_range": used_range,
        "max_row": worksheet.max_row,
        "max_column": worksheet.max_column,
    }


def sheet_query(
    store: WorkbookStore,
    *,
    workbook_id: str,
    sheet: str,
) -> dict[str, object]:
    session = store.get(workbook_id)
    worksheet = get_worksheet(session.workbook, sheet)

    return {
        "workbook_id": workbook_id,
        "name": worksheet.title,
        "position": session.workbook.worksheets.index(worksheet) + 1,
        "visibility": "very_hidden" if worksheet.sheet_state == "veryHidden" else worksheet.sheet_state,
        "is_active": session.workbook.active is worksheet,
        "max_row": worksheet.max_row,
        "max_column": worksheet.max_column,
        "sheet_dimensions": str(worksheet.sheet_dimensions),
    }
