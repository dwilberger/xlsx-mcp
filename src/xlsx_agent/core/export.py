from __future__ import annotations

import csv
import json
from io import StringIO
from pathlib import Path

from openpyxl.utils import range_boundaries

from xlsx_agent.core.errors import XlsxAgentError
from xlsx_agent.core.sheets import get_worksheet
from xlsx_agent.core.store import WorkbookStore


def _cell_to_value(cell) -> object:
    v = cell.value
    if v is None:
        return None
    if isinstance(v, bool):
        return v
    if isinstance(v, (int, float)):
        return v
    return str(v)


def sheet_to_csv(
    store: WorkbookStore,
    *,
    workbook_id: str,
    sheet: str,
    delimiter: str = ",",
) -> dict[str, object]:
    session = store.get(workbook_id)
    worksheet = get_worksheet(session.workbook, sheet)

    output = StringIO()
    writer = csv.writer(output, delimiter=delimiter)

    for row in worksheet.iter_rows():
        writer.writerow([_cell_to_value(cell) for cell in row])

    csv_content = output.getvalue()
    output.close()

    return {
        "workbook_id": workbook_id,
        "sheet": sheet,
        "csv": csv_content,
        "row_count": worksheet.max_row,
        "column_count": worksheet.max_column,
    }


def range_to_csv(
    store: WorkbookStore,
    *,
    workbook_id: str,
    sheet: str,
    range: str,
    delimiter: str = ",",
) -> dict[str, object]:
    session = store.get(workbook_id)
    worksheet = get_worksheet(session.workbook, sheet)

    try:
        min_col, min_row, max_col, max_row, resolved = range_boundaries(range)
    except ValueError as exc:
        raise XlsxAgentError(
            code="invalid_range",
            message=f"Range '{range}' is invalid.",
        ) from exc

    output = StringIO()
    writer = csv.writer(output, delimiter=delimiter)

    for row in worksheet.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        writer.writerow([_cell_to_value(cell) for cell in row])

    csv_content = output.getvalue()
    output.close()

    return {
        "workbook_id": workbook_id,
        "sheet": sheet,
        "range": resolved,
        "csv": csv_content,
        "row_count": max_row - min_row + 1,
        "column_count": max_col - min_col + 1,
    }


def csv_to_sheet(
    store: WorkbookStore,
    *,
    workbook_id: str,
    sheet: str,
    csv_data: str,
    start_cell: str = "A1",
    delimiter: str = ",",
) -> dict[str, object]:
    session = store.get(workbook_id)
    if session.read_only:
        raise XlsxAgentError(
            code="unsupported_operation",
            message=f"Workbook '{workbook_id}' was opened in read-only mode.",
        )

    worksheet = get_worksheet(session.workbook, sheet)

    from openpyxl.utils.cell import coordinate_to_tuple
    try:
        start_row, start_col = coordinate_to_tuple(start_cell)
    except ValueError as exc:
        raise XlsxAgentError(
            code="invalid_input",
            message=f"Cell address '{start_cell}' is invalid.",
        ) from exc

    reader = csv.reader(StringIO(csv_data), delimiter=delimiter)
    rows_written = 0
    cols_written = 0

    for r_idx, row in enumerate(reader):
        for c_idx, value in enumerate(row):
            cell = worksheet.cell(row=start_row + r_idx, column=start_col + c_idx)
            try:
                cell.value = float(value) if value else None
            except ValueError:
                cell.value = value
        rows_written += 1
        cols_written = max(cols_written, len(row))

    session.dirty = True
    return {
        "workbook_id": workbook_id,
        "sheet": sheet,
        "start_cell": start_cell.upper(),
        "rows_written": rows_written,
        "columns_written": cols_written,
        "dirty": session.dirty,
    }


def sheet_to_json(
    store: WorkbookStore,
    *,
    workbook_id: str,
    sheet: str,
    use_header_row: bool = True,
) -> dict[str, object]:
    session = store.get(workbook_id)
    worksheet = get_worksheet(session.workbook, sheet)

    rows = []
    headers = None

    for row_idx, row in enumerate(worksheet.iter_rows()):
        values = [_cell_to_value(cell) for cell in row]
        if row_idx == 0 and use_header_row:
            headers = [str(v) if v is not None else f"col_{i}" for i, v in enumerate(values)]
        elif headers:
            rows.append(dict(zip(headers, values, strict=False)))
        else:
            rows.append(values)

    return {
        "workbook_id": workbook_id,
        "sheet": sheet,
        "data": rows,
        "row_count": len(rows),
    }


def json_to_sheet(
    store: WorkbookStore,
    *,
    workbook_id: str,
    sheet: str,
    data: list[dict[str, object] | list[object]],
    start_cell: str = "A1",
) -> dict[str, object]:
    session = store.get(workbook_id)
    if session.read_only:
        raise XlsxAgentError(
            code="unsupported_operation",
            message=f"Workbook '{workbook_id}' was opened in read-only mode.",
        )

    worksheet = get_worksheet(session.workbook, sheet)

    from openpyxl.utils.cell import coordinate_to_tuple
    try:
        start_row, start_col = coordinate_to_tuple(start_cell)
    except ValueError as exc:
        raise XlsxAgentError(
            code="invalid_input",
            message=f"Cell address '{start_cell}' is invalid.",
        ) from exc

    if not data:
        return {
            "workbook_id": workbook_id,
            "sheet": sheet,
            "start_cell": start_cell.upper(),
            "rows_written": 0,
            "dirty": session.dirty,
        }

    first = data[0]
    if isinstance(first, dict):
        headers = list(first.keys())
        for c_idx, key in enumerate(headers):
            worksheet.cell(row=start_row, column=start_col + c_idx).value = key

        for r_idx, row in enumerate(data):
            if not isinstance(row, dict):
                raise XlsxAgentError(
                    code="invalid_input",
                    message="All rows must use the same shape.",
                )
            for c_idx, key in enumerate(headers):
                worksheet.cell(row=start_row + 1 + r_idx, column=start_col + c_idx).value = row.get(key)

        rows_written = len(data) + 1
    else:
        for r_idx, row in enumerate(data):
            if isinstance(row, dict):
                raise XlsxAgentError(
                    code="invalid_input",
                    message="All rows must use the same shape.",
                )
            for c_idx, value in enumerate(row):
                worksheet.cell(row=start_row + r_idx, column=start_col + c_idx).value = value

        rows_written = len(data)

    session.dirty = True
    return {
        "workbook_id": workbook_id,
        "sheet": sheet,
        "start_cell": start_cell.upper(),
        "rows_written": rows_written,
        "dirty": session.dirty,
    }
