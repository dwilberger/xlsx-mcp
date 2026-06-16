from __future__ import annotations

from openpyxl.utils import range_boundaries

from xlsx_agent.core.errors import XlsxAgentError
from xlsx_agent.core.sheets import get_worksheet
from xlsx_agent.core.store import WorkbookStore


def range_style_set(
    store: WorkbookStore,
    *,
    workbook_id: str,
    sheet: str,
    range: str,
    font_name: str | None = None,
    font_size: float | None = None,
    bold: bool | None = None,
    italic: bool | None = None,
    underline: str | None = None,
    fill_color: str | None = None,
    font_color: str | None = None,
    number_format: str | None = None,
    alignment: str | None = None,
) -> dict[str, object]:
    session = store.get(workbook_id)
    if session.read_only:
        raise XlsxAgentError(
            code="unsupported_operation",
            message=f"Workbook '{workbook_id}' was opened in read-only mode.",
        )

    worksheet = get_worksheet(session.workbook, sheet)

    try:
        min_col, min_row, max_col, max_row, resolved = range_boundaries(range)
    except ValueError as exc:
        raise XlsxAgentError(
            code="invalid_range",
            message=f"Range '{range}' is invalid.",
        ) from exc

    from openpyxl.styles import Alignment, Font, PatternFill

    count = 0
    for row in worksheet.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            if font_name or font_size is not None or bold is not None or italic is not None or underline or font_color:
                font = cell.font.copy()
                if font_name:
                    font = Font(name=font_name, size=font.size, bold=font.bold, italic=font.italic, color=font.color)
                if font_size is not None:
                    font = Font(name=font.name, size=font_size, bold=font.bold, italic=font.italic, color=font.color)
                if bold is not None:
                    font = Font(name=font.name, size=font.size, bold=bold, italic=font.italic, color=font.color)
                if italic is not None:
                    font = Font(name=font.name, size=font.size, bold=font.bold, italic=italic, color=font.color)
                if underline:
                    font = Font(name=font.name, size=font.size, bold=font.bold, italic=font.italic, underline=underline, color=font.color)
                if font_color:
                    font = Font(name=font.name, size=font.size, bold=font.bold, italic=font.italic, color=font_color)
                cell.font = font

            if fill_color:
                cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")

            if number_format:
                cell.number_format = number_format

            if alignment:
                cell.alignment = Alignment(horizontal=alignment)

            count += 1

    session.dirty = True
    return {
        "workbook_id": workbook_id,
        "sheet": sheet,
        "range": resolved,
        "styled_count": count,
        "dirty": session.dirty,
    }


def conditional_formatting_list(
    store: WorkbookStore,
    *,
    workbook_id: str,
    sheet: str,
) -> dict[str, object]:
    session = store.get(workbook_id)
    worksheet = get_worksheet(session.workbook, sheet)

    rules = []
    for cf in worksheet.conditional_formatting:
        for rule in cf.rules:
            rules.append({
                "type": rule.type,
                "priority": rule.priority,
                "ranges": str(cf.sqref) if cf.sqref else None,
                "formula": getattr(rule, "formula", None),
            })

    return {
        "workbook_id": workbook_id,
        "sheet": sheet,
        "rules": rules,
        "count": len(rules),
    }


def conditional_formatting_add(
    store: WorkbookStore,
    *,
    workbook_id: str,
    sheet: str,
    range: str,
    type: str,
    formula: str | None = None,
    priority: int = 1,
) -> dict[str, object]:
    session = store.get(workbook_id)
    if session.read_only:
        raise XlsxAgentError(
            code="unsupported_operation",
            message=f"Workbook '{workbook_id}' was opened in read-only mode.",
        )

    worksheet = get_worksheet(session.workbook, sheet)

    from openpyxl.formatting.rule import CellIsRule, FormulaRule
    if type == "cellIs":
        rule = CellIsRule(operator="equal", formula=[formula] if formula else None)
    elif type == "formula":
        rule = FormulaRule(formula=[formula] if formula else None)
    else:
        rule = CellIsRule(operator=type)

    worksheet.conditional_formatting.add(range, rule)

    session.dirty = True
    return {
        "workbook_id": workbook_id,
        "sheet": sheet,
        "range": range,
        "type": type,
        "dirty": session.dirty,
    }


def conditional_formatting_remove(
    store: WorkbookStore,
    *,
    workbook_id: str,
    sheet: str,
    range: str,
) -> dict[str, object]:
    session = store.get(workbook_id)
    if session.read_only:
        raise XlsxAgentError(
            code="unsupported_operation",
            message=f"Workbook '{workbook_id}' was opened in read-only mode.",
        )

    worksheet = get_worksheet(session.workbook, sheet)

    removed = 0
    for cf in list(worksheet.conditional_formatting):
        if str(cf.sqref) == range:
            worksheet.conditional_formatting.remove(cf)
            removed += 1

    session.dirty = True
    return {
        "workbook_id": workbook_id,
        "sheet": sheet,
        "removed_count": removed,
        "dirty": session.dirty,
    }
