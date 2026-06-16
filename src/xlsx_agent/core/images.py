from __future__ import annotations

from pathlib import Path

from openpyxl.drawing.image import Image as OpenpyxlImage
from openpyxl.utils import get_column_letter

from xlsx_agent.core.errors import XlsxAgentError
from xlsx_agent.core.sheets import get_worksheet
from xlsx_agent.core.store import WorkbookStore


def image_add(
    store: WorkbookStore,
    *,
    workbook_id: str,
    sheet: str,
    path: str,
    anchor: str,
    width: int | None = None,
    height: int | None = None,
) -> dict[str, object]:
    session = store.get(workbook_id)
    if session.read_only:
        raise XlsxAgentError(
            code="unsupported_operation",
            message=f"Workbook '{workbook_id}' was opened in read-only mode.",
        )

    worksheet = get_worksheet(session.workbook, sheet)

    image_path = Path(path)
    if not image_path.is_file():
        raise XlsxAgentError(
            code="file_not_found",
            message=f"Image file '{path}' was not found.",
        )

    try:
        img = OpenpyxlImage(str(image_path))
    except Exception as exc:
        raise XlsxAgentError(
            code="invalid_input",
            message=f"Failed to load image file '{path}'.",
            details={"error": str(exc)},
        ) from exc

    if width is not None:
        img.width = width
    if height is not None:
        img.height = height

    img.anchor = anchor
    worksheet.add_image(img)

    session.dirty = True
    return {
        "workbook_id": workbook_id,
        "sheet": sheet,
        "path": str(image_path),
        "anchor": anchor,
        "width": img.width,
        "height": img.height,
        "dirty": session.dirty,
    }


def image_list(
    store: WorkbookStore,
    *,
    workbook_id: str,
    sheet: str,
) -> dict[str, object]:
    session = store.get(workbook_id)
    worksheet = get_worksheet(session.workbook, sheet)

    images = []
    for idx, img in enumerate(worksheet._images):
        anchor = img.anchor
        if hasattr(anchor, "_from"):
            from_cell = f"{get_column_letter(anchor._from.col + 1)}{anchor._from.row + 1}"
        else:
            from_cell = str(anchor)

        images.append({
            "index": idx,
            "anchor": from_cell,
            "width": img.width,
            "height": img.height,
        })

    return {
        "workbook_id": workbook_id,
        "sheet": sheet,
        "images": images,
        "count": len(images),
    }


def image_remove(
    store: WorkbookStore,
    *,
    workbook_id: str,
    sheet: str,
    index: int,
) -> dict[str, object]:
    session = store.get(workbook_id)
    if session.read_only:
        raise XlsxAgentError(
            code="unsupported_operation",
            message=f"Workbook '{workbook_id}' was opened in read-only mode.",
        )

    worksheet = get_worksheet(session.workbook, sheet)

    if index < 0 or index >= len(worksheet._images):
        raise XlsxAgentError(
            code="invalid_input",
            message=f"Image index {index} is out of range.",
            details={"available_count": len(worksheet._images)},
        )

    removed_img = worksheet._images[index]
    anchor = removed_img.anchor
    if hasattr(anchor, "_from"):
        from_cell = f"{get_column_letter(anchor._from.col + 1)}{anchor._from.row + 1}"
    else:
        from_cell = str(anchor)

    del worksheet._images[index]

    session.dirty = True
    return {
        "workbook_id": workbook_id,
        "sheet": sheet,
        "removed_index": index,
        "anchor": from_cell,
        "dirty": session.dirty,
    }
